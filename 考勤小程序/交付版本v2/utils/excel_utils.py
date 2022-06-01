import shutil

import numpy as np
from openpyxl import load_workbook
from openpyxl.comments import Comment
import traceback
import pandas as pd
import pyttsx3
import datetime
import json
import os
import re


def read_excel_pandas(file_path):
    """读取数据，获取人员名单"""
    try:
        data = pd.read_excel(file_path)
        return data
    except ValueError:
        traceback.print_exc()
        raise Exception("您的excel表格中可能没有数据，请您核验一下。")


def read_excel_cs(file_path):
    """ 获取班级和学生名单 """
    sheets = get_excel_sheets(file_path)
    course_info = {}  # 课程的数据（包括班级、学生名单）
    for s in sheets:
        try:
            data = pd.read_excel(file_path, sheet_name=s)
        except ValueError:
            traceback.print_exc()
            raise Exception("您的excel表格中可能没有数据，请您核验一下。")
        cla, stu, cs_dict = data['班级'], data['名单'], {}
        # 先写成这种格式：{"一班":[], "二班":[], "三班":[]}
        for c in set(cla):
            cs_dict[c] = []
        # 对data进行遍历， 将学生名单插入对应对班级
        for i in range(len(data)):
            row = data.iloc[i]
            cs_dict[row['班级']].append(row['名单'])
        course_info[s] = cs_dict
    return course_info


def save_excel_pandas(data, save_path):
    """将考勤结果保存到excel"""
    data.to_excel(save_path, index=False)
    input(f'考勤结果已经保存至【{save_path}】, 请您查收。')


def modify_comment(com_str, read_path, save_path=None):
    """添加、修改excel的备注"""
    # todo
    # wb = load_workbook(filename=read_path)
    # ws = wb.active
    # comment = Comment(com_str, 'cpvs')
    # ws['A1'].comment = comment
    #
    # if not save_path:
    #     save_path = read_path
    # wb.save(save_path)


def get_excel_sheets(file_path):
    """ 获取excel的sheet """
    wb = load_workbook(filename=file_path)
    return wb.sheetnames


def read_excel_openpyxl(file_path):
    """使用openpyxl包读取excel"""
    # wb = load_workbook(filename=file_path)
    # # todo
    # # 获取所有的sheet名
    # sheets = wb.sheetnames
    # print(sheets)
    # all_table = wb.worksheets
    # print(all_table)
    # aaa = wb["总表"]
    # print(aaa['A1'].value)
    # print(aaa['A2'].value)
    # print(aaa['A3'].value)
    # for i in range(1, aaa.max_row):
    #     print(aaa.cell(row=i, column=2).value)


def save_excel_openpyxl(data, save_path):
    """ 使用openpyxl保存数据到excel """
    # todo


if __name__ == '__main__':
    file = "../1-学生名单表/学生名单.xlsx"
    # print(get_excel_sheets(file))
    # wb = load_workbook(filename=file)
    # read_sheet(wb, '语文')
    print(read_excel_cs(file))
