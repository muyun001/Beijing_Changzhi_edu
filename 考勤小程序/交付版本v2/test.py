from openpyxl import load_workbook

path_file = r"1-学生名单表/工作簿1.xlsx"
wb = load_workbook(filename=path_file)

# 获取所有的sheet名
sheets = wb.sheetnames
print(sheets)
all_table = wb.worksheets
print(all_table)
aaa = wb["总表"]
print(aaa['A1'].value)
print(aaa['A2'].value)
print(aaa['A3'].value)
for i in range(1, aaa.max_row):
    print(aaa.cell(row=i, column=2).value)
