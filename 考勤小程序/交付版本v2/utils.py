from openpyxl import load_workbook
from openpyxl.comments import Comment
import traceback
import pandas as pd
import pyttsx3
import datetime
import json
import os
import re


def get_files(folder_path):
    """获取文件夹下的所有文件"""
    return os.listdir(folder_path)


def read_excel(file_path):
    """读取数据，获取人员名单"""
    try:
        data = pd.read_excel(file_path)
        return data
    except ValueError:
        traceback.print_exc()
        raise Exception("您的excel表格中可能没有数据，请您核验一下。")


def voice_read():
    engine = pyttsx3.init()
    engine.setProperty('rate', 150)  # 设置语速
    engine.setProperty('volume', 2.0)  # 设置音量
    return engine


def get_date():
    """获取当前的年月日"""
    # 获取当地时间
    now = datetime.datetime.now()
    return now.year, now.month, now.day


def save_excel(data, save_path):
    """将考勤结果保存到excel"""
    data.to_excel(save_path, index=False)
    input(f'考勤结果已经保存至【{save_path}】, 请您查收。')


def lesson_judge():
    """ 判断此时是第几节课 """
    # 每节课的时间段
    lessons = [
        ["08:15:00", "08:55:00"],  # 第1节课
        ["09:10:00", "09:50:00"],  # 第2节课
        ["10:25:00", "11:05:00"],  # 第3节课
        ["11:20:00", "12:00:00"],  # 第4节课

        ["13:30:00", "14:10:00"],  # 第5节课
        ["14:25:00", "15:05:00"],  # 第6节课
        ["15:25:00", "16:05:00"],  # 第7节课
    ]
    now = datetime.datetime.now()
    today = str(datetime.date.today())

    # 判断此时是第几节课
    index = 0
    for time_zone in lessons:
        index += 1
        start_time = datetime.datetime.strptime(f"{today} {time_zone[0]}", "%Y-%m-%d %H:%M:%S")
        end_time = datetime.datetime.strptime(f"{today} {time_zone[1]}", "%Y-%m-%d %H:%M:%S")

        # todo 需要完善
        if start_time < now < end_time:
            return index
    else:
        print("现在好像是课余时间")


def read_settings():
    """读取配置文件"""
    setting_path = r"./0-说明文档和配置文档/配置文档.txt"
    with open(setting_path, "r", encoding="utf-8") as f:
        text = f.read()
        # 正则表达式：re
        # 读取"是否需要开启程序语音"
        is_open_voice = re.search(".*语音播报.*?【(\d)】", text, re.S)
        read_folder = re.search("需要考勤的目录名：【(.*?)】", text, re.S)
        save_folder = re.search("生成考勤结果的目录名：【(.*?)】", text, re.S)
        kq_abnormal = re.search("""考勤异常情况：(.*?)注意: .*【考勤异常】""", text, re.S)
        if not is_open_voice:
            raise Exception("是否需要语音播报？需要您在【配置文件.txt】中进行配置。")
        if not read_folder.group(1):
            raise Exception("考勤的目录名，需要您在【配置文件.txt】中进行配置。")
        if not save_folder.group(1):
            raise Exception("保存考勤结果的目录名，需要您在【配置文件.txt】中进行配置。")

        return int(is_open_voice.group(1)), \
               read_folder.group(1), \
               save_folder.group(1), \
               kq_abnormal.group(1)


def dict_to_str(dic):
    """
    将字典按照一定的格式转成字符串
    dict: {"1":"迟到", "2":"请假"}
    str: 1迟到，2请假
    """
    dic_str = json.dumps(dic, separators=('，', ''), ensure_ascii=False)
    return dic_str.replace('"', "").lstrip("{").rstrip("}")


def str_to_dict(s):
    """
    将一定格式的字符串转成字典
    str: '\n1、请假\n2、迟到\n3、早退\n4、旷课\n5、做核酸\n6、备赛\n7、参加学校活动/任务\n8、其他情况\n'
    dict: {'1': '请假', '2': '迟到', '3': '早退', '4': '旷课', '5': '做核酸', '6': '备赛', '7': '参加学校活动/任务', '8': '其他情况'}
    """
    sit_dict = {}
    for i in s.strip("\n").split('\n'):
        if not i.strip():
            continue
        j = i.split('、')
        sit_dict[j[0]] = j[1]
    return sit_dict


def key_info(s):
    """
    获取文件名的关键信息
    s: 2班学生名单.xlsx
    return: 2班
    """
    return s.replace(".xlsx", "") \
        .replace("名单", "") \
        .replace("学生", "") \
        .replace("昌", "") \
        .replace("平", "") \
        .replace("职", "") \
        .replace("北京", "")


def modify_comment(com_str, read_path, save_path=None):
    """添加、修改excel的备注"""
    # todo
    wb = load_workbook(filename=read_path)
    ws = wb.active
    comment = Comment(com_str, 'yunan')
    ws['A1'].comment = comment

    if not save_path:
        save_path = read_path
    wb.save(save_path)


if __name__ == '__main__':
    print(read_settings())
    # print(kq_files())
    # path = "./0-说明文档和配置文档"
    # print(get_files(path))
