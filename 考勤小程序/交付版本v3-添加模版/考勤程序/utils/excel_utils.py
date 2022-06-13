import os.path
import openpyxl
from openpyxl import load_workbook, Workbook
import traceback
import pandas as pd
import date_utils


def read_excel_pandas(file_path):
    """读取数据，获取人员名单"""
    try:
        data = pd.read_excel(file_path)
        return data
    except ValueError:
        traceback.print_exc()
        raise Exception("您的excel表格中可能没有数据，请您核验一下。")


def read_excel_cs(file_path):
    """ 获取科目、班级和学生名单 """
    sheets = get_excel_sheets(file_path)
    course_info = {}  # 课程的数据（包括班级、学生名单）
    for s in sheets:
        try:
            data = pd.read_excel(file_path, sheet_name=s)
        except ValueError:
            traceback.print_exc()
            raise Exception("您的excel表格中可能没有数据，请您核验一下。")
        cla, stu, cs_dict = data['班级'], data['名单'], {}
        # 先写成这种格式：{"一班":[], "二班":[], "三班":[]}
        for c in set(cla):
            cs_dict[c] = []
        # 对data进行遍历， 将学生名单插入对应对班级
        for i in range(len(data)):
            row = data.iloc[i]
            cs_dict[row['班级']].append(row['名单'])
        course_info[s] = cs_dict
    return course_info


def save_excel_pandas(data, save_path):
    """将考勤结果保存到excel"""
    # todo 不好用，暂时弃用
    d = pd.read_excel(save_path, sheet_name=data['course'])
    kq = []
    for i in range(len(d)):
        row = d.iloc[i]
        cls = row['班级']
        s = row['名单']
        if cls == data['class']:
            key = f"kq-{cls}-{s}"
            kq.append(data[key])
        else:
            kq.append("")
    d["考勤"] = kq

    writer = pd.ExcelWriter(save_path)
    d.to_excel(writer, sheet_name=data['course'], index=False)
    writer.save()


def modify_comment(com_str, read_path, save_path=None):
    """添加、修改excel的备注"""
    # todo
    # wb = load_workbook(filename=read_path)
    # ws = wb.active
    # comment = Comment(com_str, 'cpvs')
    # ws['A1'].comment = comment
    #
    # if not save_path:
    #     save_path = read_path
    # wb.save(save_path)


def get_excel_sheets(file_path):
    """ 获取excel的sheet """
    wb = load_workbook(filename=file_path)
    return wb.sheetnames


def read_excel_openpyxl(file_path):
    """使用openpyxl包读取excel"""
    # wb = load_workbook(filename=file_path)
    # # todo
    # # 获取所有的sheet名
    # sheets = wb.sheetnames
    # print(sheets)
    # all_table = wb.worksheets
    # print(all_table)
    # aaa = wb["总表"]
    # print(aaa['A1'].value)
    # print(aaa['A2'].value)
    # print(aaa['A3'].value)
    # for i in range(1, aaa.max_row):
    #     print(aaa.cell(row=i, column=2).value)


def save_excel_openpyxl(data, save_path):
    """ 使用openpyxl保存数据到excel """
    from openpyxl.styles import Font
    try:
        wb = openpyxl.load_workbook(save_path)
    except Exception as e:
        print(e)
    ws = wb[data["course"]]
    ws.insert_cols(3)  # 将新的一列插入到第3列

    list_ws = list(ws)
    new_column = list(ws['C'])
    # 设置标题（20220608周五-L2）
    title = "{today}{weekday}-{lesson}".format(
        today=date_utils.get_date(),
        weekday=date_utils.get_weekday(),
        lesson=data['lesson']
    )
    new_column[0].value = title

    for i in range(1, ws.max_row):
        excel_cls = list_ws[i][0].value  # 班级
        excel_s = list_ws[i][1].value  # 学生
        if excel_cls in list(data["考勤情况"].keys()):
            kq = data["考勤情况"][excel_cls][excel_s]  # 考勤情况
            new_column[i].value = kq
            if kq != "考勤正常":
                new_column[i].font = Font(color="FF0000")  # 设置字体颜色为红色

    wb.save(save_path)


if __name__ == '__main__':
    data = {'course': '语文', '考勤情况': {
        '2班': {'侯文杰': '早退', '刘子豪': '早退', '刘洋': '早退', '吕铭': '做核酸', '杨悦昕': '做核酸', '杨智恒': '考勤正常', '杨梓琦': '考勤正常',
               '白轶臣': '早退', '蒋文拓': '早退', '赵梓豪': '早退', '靳馨娴': '考勤正常', '黄新之': '备赛'}}}
    path = os.path.abspath('../..') + "/2-考勤结果/考勤总表.xlsx"
    # save_excel_pandas(data, path)
    print(path)
    save_excel_openpyxl(data, path)
